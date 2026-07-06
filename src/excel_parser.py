import logging
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from .config_loader import ApiConfig, ExcelConfig
from .image_extractor import extract_images_by_row

logger = logging.getLogger(__name__)


def parse_excel(
    excel_path: str,
    excel_config: ExcelConfig,
    api_config: ApiConfig,  # 현재 미사용, 확장을 위해 유지
) -> List[Dict[str, Any]]:
    """
    Excel 파일을 파싱하여 API 전송용 행 데이터 목록을 반환한다.

    반환 형식:
        [
          {
            "deviceId": "...",
            "wfProcess": "...",
            ...
            "files": [("image1.png", b"..."), ...]
          },
          ...
        ]
    """
    sheet_name: Optional[str] = excel_config.sheet_name
    text_columns = excel_config.text_columns

    # 1) 이미지 추출 (row 번호 기준 매핑)
    logger.info("Excel 이미지 추출 시작...")
    images_by_row = extract_images_by_row(excel_path, sheet_name)

    # 2) 셀 데이터 읽기
    logger.info("Excel 셀 데이터 읽기 시작...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        logger.warning("Excel 파일에 데이터가 없습니다.")
        return []

    # 첫 번째 행 = 헤더
    header_map: Dict[str, int] = {
        str(cell).strip() if cell is not None else "": idx
        for idx, cell in enumerate(all_rows[0])
    }
    logger.info(f"감지된 컬럼: {[k for k in header_map if k]}")

    # 설정된 컬럼이 헤더에 존재하는지 확인
    for col in text_columns:
        if col.excel_name not in header_map:
            logger.warning(
                f"컬럼 '{col.excel_name}'을 Excel에서 찾을 수 없습니다. "
                f"존재하는 컬럼: {list(header_map.keys())}"
            )

    # 3) 2행(index=1)부터 데이터 행 처리
    parsed: List[Dict[str, Any]] = []

    for excel_row_num, row_values in enumerate(all_rows[1:], start=2):
        row_dict: Dict[str, Any] = {}

        for col in text_columns:
            col_idx = header_map.get(col.excel_name)
            if col_idx is not None and col_idx < len(row_values):
                raw = row_values[col_idx]
                row_dict[col.dto_field] = str(raw) if raw is not None else ""
            else:
                row_dict[col.dto_field] = ""

        row_dict["files"] = images_by_row.get(excel_row_num, [])

        if _is_empty_row(row_dict, [c.dto_field for c in text_columns]):
            logger.debug(f"행 {excel_row_num}: 빈 행 스킵")
            continue

        parsed.append(row_dict)
        logger.info(
            f"행 {excel_row_num} 파싱 완료 | "
            f"이미지 {len(row_dict['files'])}개"
        )

    logger.info(f"총 {len(parsed)}개 행 파싱 완료")
    return parsed


def _is_empty_row(row_dict: Dict[str, Any], dto_fields: List[str]) -> bool:
    if not dto_fields:
        return False
    for field in dto_fields:
        value = row_dict.get(field, "")
        if isinstance(value, list):
            continue  # list(files 등)는 텍스트 빈값 판단에서 제외
        if str(value).strip():
            return False
    return True
