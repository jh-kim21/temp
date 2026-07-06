"""
Excel(.xlsx) 파일 내부 이미지를 추출하여 행 번호(1-based)에 매핑한다.

xlsx는 ZIP 아카이브이므로, openpyxl 내부 표현에 의존하지 않고
직접 ZIP을 열어 XML 파싱으로 anchor 위치와 이미지 데이터를 가져온다.
이를 통해 openpyxl 버전 차이에 무관하게 동작한다.
"""

import logging
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# XML 네임스페이스 상수
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_OFF_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_SPREADSHEET = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"


def extract_images_by_row(
    excel_path: str, sheet_name: Optional[str] = None
) -> Dict[int, List[Tuple[str, bytes]]]:
    """
    Excel 파일에서 이미지를 추출하고 행 번호(1-based)로 매핑한다.

    Returns:
        {row_number: [(filename, image_bytes), ...]}
    """
    sheet_idx = _resolve_sheet_index(excel_path, sheet_name)
    result = _extract_via_zip(excel_path, sheet_idx)

    total = sum(len(imgs) for imgs in result.values())
    logger.info(f"이미지 추출 완료: {total}개 / {len(result)}개 행")
    return result


def _resolve_sheet_index(excel_path: str, sheet_name: Optional[str]) -> int:
    wb = load_workbook(excel_path, read_only=True)
    try:
        if not sheet_name:
            return 0
        if sheet_name not in wb.sheetnames:
            logger.warning(f"시트 '{sheet_name}'를 찾을 수 없어 첫 번째 시트를 사용합니다.")
            return 0
        return wb.sheetnames.index(sheet_name)
    finally:
        wb.close()


def _extract_via_zip(
    excel_path: str, sheet_idx: int
) -> Dict[int, List[Tuple[str, bytes]]]:
    result: Dict[int, List[Tuple[str, bytes]]] = {}

    with zipfile.ZipFile(excel_path, "r") as z:
        namelist = set(z.namelist())

        sheet_rels_path = _get_sheet_rels_path(z, namelist, sheet_idx)
        if not sheet_rels_path:
            logger.info(f"시트 {sheet_idx}의 관계 파일을 찾을 수 없습니다.")
            return result

        drawing_path = _find_drawing_path(z, sheet_rels_path)
        if not drawing_path:
            logger.info("Drawing 파일이 없습니다 (이미지 없음).")
            return result

        logger.debug(f"Drawing 파일: {drawing_path}")

        rId_to_media = _parse_drawing_rels(z, namelist, drawing_path)
        if not rId_to_media:
            logger.info("Drawing 관계 파일에 미디어가 없습니다.")
            return result

        with z.open(drawing_path) as f:
            root = ET.parse(f).getroot()

        for anchor in root:
            row = _get_anchor_row(anchor)
            if row is None:
                continue

            for pic in anchor.iter(f"{{{_NS_XDR}}}pic"):
                img_bytes, filename = _read_pic_image(z, namelist, pic, rId_to_media)
                if img_bytes:
                    result.setdefault(row, []).append((filename, img_bytes))
                    logger.debug(f"행 {row}: 이미지 '{filename}' 추출")

    return result


def _get_sheet_rels_path(
    z: zipfile.ZipFile, namelist: set, sheet_idx: int
) -> Optional[str]:
    """워크북 관계를 파싱하여 sheet_idx에 해당하는 시트의 _rels 경로를 반환한다."""
    # 가장 흔한 경우: sheet1.xml, sheet2.xml 순서
    direct = f"xl/worksheets/_rels/sheet{sheet_idx + 1}.xml.rels"
    if direct in namelist:
        return direct

    # 시트 순서가 다를 경우: workbook.xml.rels 파싱으로 정확한 파일명 확인
    wb_rels = "xl/_rels/workbook.xml.rels"
    wb_xml = "xl/workbook.xml"
    if wb_rels not in namelist or wb_xml not in namelist:
        return None

    with z.open(wb_rels) as f:
        rels_root = ET.parse(f).getroot()

    rId_to_target: Dict[str, str] = {}
    for rel in rels_root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        if "worksheet" in rel.get("Type", "").lower():
            rId_to_target[rel.get("Id", "")] = rel.get("Target", "")

    with z.open(wb_xml) as f:
        wb_root = ET.parse(f).getroot()

    sheet_targets: List[str] = []
    for sheet in wb_root.iter(f"{{{_NS_SPREADSHEET}}}sheet"):
        r_id = sheet.get(f"{{{_NS_OFF_REL}}}id", "")
        if r_id in rId_to_target:
            sheet_targets.append(rId_to_target[r_id])

    if sheet_idx >= len(sheet_targets):
        return None

    # 'worksheets/sheet2.xml' -> 'xl/worksheets/_rels/sheet2.xml.rels'
    target = sheet_targets[sheet_idx].lstrip("/")
    filename = target.split("/")[-1]
    rels_path = f"xl/worksheets/_rels/{filename}.rels"
    return rels_path if rels_path in namelist else None


def _find_drawing_path(z: zipfile.ZipFile, sheet_rels_path: str) -> Optional[str]:
    """시트 관계 파일에서 Drawing XML 경로를 찾는다."""
    with z.open(sheet_rels_path) as f:
        root = ET.parse(f).getroot()

    for rel in root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        if "drawing" in rel.get("Type", "").lower():
            target = rel.get("Target", "")
            # '../drawings/drawing1.xml' -> 'xl/drawings/drawing1.xml'
            return _resolve_path("xl/worksheets", target)

    return None


def _parse_drawing_rels(
    z: zipfile.ZipFile, namelist: set, drawing_path: str
) -> Dict[str, str]:
    """Drawing 관계 파일을 파싱하여 rId -> 미디어 경로를 반환한다."""
    dir_part, file_part = drawing_path.rsplit("/", 1)
    rels_path = f"{dir_part}/_rels/{file_part}.rels"

    if rels_path not in namelist:
        return {}

    with z.open(rels_path) as f:
        root = ET.parse(f).getroot()

    rId_to_media: Dict[str, str] = {}
    for rel in root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        r_id = rel.get("Id")
        target = rel.get("Target", "")
        if r_id and target:
            # '../media/image1.png' -> 'xl/media/image1.png'
            rId_to_media[r_id] = _resolve_path(dir_part, target)

    return rId_to_media


def _get_anchor_row(anchor: ET.Element) -> Optional[int]:
    """twoCellAnchor / oneCellAnchor에서 1-based 행 번호를 추출한다."""
    tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
    if tag not in ("twoCellAnchor", "oneCellAnchor"):
        return None

    from_elem = anchor.find(f"{{{_NS_XDR}}}from")
    if from_elem is None:
        return None

    row_elem = from_elem.find(f"{{{_NS_XDR}}}row")
    if row_elem is None or not row_elem.text:
        return None

    try:
        return int(row_elem.text) + 1  # 0-based -> 1-based
    except ValueError:
        return None


def _read_pic_image(
    z: zipfile.ZipFile,
    namelist: set,
    pic: ET.Element,
    rId_to_media: Dict[str, str],
) -> Tuple[Optional[bytes], str]:
    """xdr:pic 요소에서 이미지 bytes와 파일명을 반환한다."""
    blip_fill = pic.find(f"{{{_NS_XDR}}}blipFill")
    if blip_fill is None:
        return None, ""

    blip = blip_fill.find(f"{{{_NS_A}}}blip")
    if blip is None:
        return None, ""

    r_embed = blip.get(f"{{{_NS_OFF_REL}}}embed")
    if not r_embed or r_embed not in rId_to_media:
        return None, ""

    media_path = rId_to_media[r_embed]
    if media_path not in namelist:
        logger.warning(f"미디어 파일이 아카이브에 없습니다: {media_path}")
        return None, ""

    with z.open(media_path) as f:
        img_bytes = f.read()

    return img_bytes, Path(media_path).name


def _resolve_path(base_dir: str, target: str) -> str:
    """상대 경로(target)를 base_dir 기준으로 절대 경로로 변환한다."""
    parts = base_dir.split("/")
    for segment in target.split("/"):
        if segment == "..":
            if parts:
                parts.pop()
        elif segment and segment != ".":
            parts.append(segment)
    return "/".join(parts)


def detect_mime_type(img_bytes: bytes) -> str:
    """이미지 매직 바이트로 MIME 타입을 감지한다."""
    if img_bytes[:4] == b"\x89PNG":
        return "image/png"
    if img_bytes[:2] == b"\xff\xd8":
        return "image/jpeg"
    if img_bytes[:4] == b"GIF8":
        return "image/gif"
    if img_bytes[:2] == b"BM":
        return "image/bmp"
    if len(img_bytes) >= 12 and img_bytes[8:12] == b"WEBP":
        return "image/webp"
    return "application/octet-stream"
